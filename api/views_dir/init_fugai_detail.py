import datetime, time
from openpyxl import Workbook
from xiongzhanghao import models
from xiongzhanghao.publicFunc import Response
from xiongzhanghao.publicFunc import account
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.db.models import Q
from backend.selectDeleteQuery import index


# 统计覆盖报表详情
@csrf_exempt
@account.is_token(models.xzh_userprofile)
def statisticalReports(request):
    response = Response.ResponseObj()
    objs = models.xzh_fugai_baobiao.objects.filter(user__role=61)
    print(objs)
    data_lsit = []
    for obj in objs:

        print('=------=-=-=-=-==-=--=-=-==-=-obj.user_id=-----> ', obj.user_id)
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="客户名称")
        ws.cell(row=1, column=2, value="关键词")
        ws.cell(row=1, column=3, value="链接")
        ws.cell(row=1, column=4, value="排名")
        # ws.cell(row=1, column=5, value="文章链接")
        # ws.cell(row=1, column=6, value="文章排名")
        ws.cell(row=1, column=5, value="创建时间")

        # # 合并单元格        开始行      结束行       用哪列          占用哪列
        ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

        # print('设置列宽')
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        # ws.column_dimensions['F'].width = 20
        # ws.column_dimensions['G'].width = 20
        row = 2
        keywordObjs = models.xzh_keywords.objects.filter(user_id=obj.user_id)
        keywords_num = keywordObjs.count()  # 关键词数
        if keywords_num > 0:
            now_date = datetime.datetime.now()
            now = now_date.strftime("%Y-%m-%d")
            q = Q(select_date__lt=now) | Q(select_date__isnull=True)
            keyword_count = keywordObjs.filter(q).count()
            # 关键词查询 情况
            data_lsit.append({
                'user_id': obj.user_id,
                'keywords_num': keywords_num,
                'keyword_count': keyword_count
            })
            if int(keyword_count) == 0:
                print('------------------------------------------------已经全部查完----------------------------------------------')
                keywordDetail = models.xzh_keywords_detail.objects.filter(
                    xzh_keywords__user_id=obj.user_id,
                    create_date=now,
                )
                fugaiNum = keywordDetail.count()
                url_list = []
                for i in keywordDetail:
                    if i.url not in url_list:
                        url_list.append(i.url)
                    ws.cell(row=row, column=1, value="{}".format(i.xzh_keywords.user.username))
                    ws.cell(row=row, column=2, value="{}".format(i.xzh_keywords.keywords))
                    ws.cell(row=row, column=3, value="{}".format(i.url))
                    ws.cell(row=row, column=4, value="{}".format(i.rank))
                    # ws.cell(row=row, column=5, value="{}".format(i.article_url))
                    # ws.cell(row=row, column=6, value="{}".format(i.article_rank))
                    ws.cell(row=row, column=5, value="{}".format(i.create_date))
                    row += 1

                urlNum = len(url_list)
                timestamp = str(time.time()) + str(obj.id) + str(time.time()).split('.')[1][2:-1]
                xlsx_url = 'statics/fugai_baobiao_xlsx/{}.xlsx'.format(obj.user.username + timestamp)
                print('xlsx_url==============================> ',xlsx_url)
                wb.save(xlsx_url)
                # baobiao_url = '127.0.0.1:8003' + '/' + xlsx_url
                baobiao_url = 'http://xiongzhanghao.zhugeyingxiao.com:8003' + '/' + xlsx_url

                fugai_baobiao_detail = models.xzh_fugai_baobiao_detail.objects.filter(xzh_fugai_baobiao_id=obj.id, create_date=now)
                data = {
                    'link_num' : urlNum,
                    'cover_num' : fugaiNum,
                    'xzh_fugai_baobiao_id' : obj.id,
                    'baobiao_url' : baobiao_url
                }
                if fugai_baobiao_detail:
                    fugai_baobiao_detail.update(**data)
                else:
                    print('===========创建')
                    fugai_baobiao_detail.create(**data)
    response.code = 200
    response.msg = '更新数据完成'
    response.data = data_lsit
    return JsonResponse(response.__dict__)






# 重新查询该用户当天覆盖
@csrf_exempt
@account.is_token(models.xzh_userprofile)
def queryAgain(request):
    response = Response.ResponseObj()
    user_id = request.GET.get('user_id')
    o_id = request.GET.get('o_id')
    userObjs = models.xzh_userprofile.objects
    role_id = userObjs.filter(id=user_id)[0].role_id
    start = datetime.datetime.now().strftime('%Y-%m-%d 00:00:00')
    stop = datetime.datetime.now().strftime('%Y-%m-%d 23:59:59')
    now =  datetime.datetime.now().strftime('%Y-%m-%d')
    print('start, stop---> ',start, stop, now)
    if int(role_id) in [64, 66]:
        user_obj = userObjs.filter(id=o_id)
        if user_obj:
            models.xzh_keywords_detail.objects.filter(create_date=now).filter(xzh_keywords__user_id=o_id).delete()
            obj = models.xzh_keywords.objects.filter(select_date__lte=stop, select_date__gte=start)
            obj.update(select_date=None)
            models.xzh_fugai_baobiao_detail.objects.filter(xzh_fugai_baobiao__user_id=o_id).filter(create_date=now).delete()
            response.code = 200
            response.msg = '重查成功'
        else:
            response.code = 301
            response.msg = '无此用户'
    else:
        response.code = 301
        response.msg = '该角色不可操作'
    return JsonResponse(response.__dict__)

















