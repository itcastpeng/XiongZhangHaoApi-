from django.shortcuts import render, HttpResponse
from xiongzhanghao import models
from xiongzhanghao.publicFunc import Response
from xiongzhanghao.publicFunc import account
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from xiongzhanghao.publicFunc.condition_com import conditionCom
from xiongzhanghao.forms.add_fans import AddForm, UpdateForm, SelectForm
from django.db.models import Q
import json, requests, datetime, time
from openpyxl import Workbook
from openpyxl.styles.alignment import Alignment
# cerf  token验证 用户展示模块
@csrf_exempt
@account.is_token(models.xzh_userprofile)
def fans(request):
    response = Response.ResponseObj()
    forms_obj = SelectForm(request.GET)
    if forms_obj.is_valid():
        current_page = forms_obj.cleaned_data['current_page']
        length = forms_obj.cleaned_data['length']
        print('forms_obj.cleaned_data -->', forms_obj.cleaned_data)
        order = request.GET.get('order', '-create_date')
        field_dict = {
            'id': '',
            'status': '',
            'username': '__contains',
            'create_date': '',
            'belong_user_id': '',
            'is_debug': 'bool',
        }
        q = conditionCom(request, field_dict)
        points = request.GET.get('points')

        print('q -->', q)
        objs = models.xzh_add_fans.objects.select_related('belong_user').filter(q).order_by(order)
        count = objs.count()
        if points:    # 查询 用户搜索条件 避免出现重复
            objs = models.xzh_add_fans.objects.select_related('belong_user').values('belong_user_id', 'belong_user__username').distinct()
            ret_data = []
            for obj in objs:
                print('obj----------> ',obj)
                ret_data.append({
                    'belong_user_id': obj.get('belong_user_id'),  # 归属人ID
                    'belong_user': obj.get('belong_user__username'),  # 归属人名字
                })
            response.code = 200
            response.data = {'ret_data':ret_data}
        else:
            if length != 0:
                start_line = (current_page - 1) * length
                stop_line = start_line + length
                objs = objs[start_line: stop_line]

            # 返回的数据
            ret_data = []
            for obj in objs:

                #  将查询出来的数据 加入列表
                ret_data.append({
                    'id': obj.id,
                    'belong_user_id':obj.belong_user_id,    # 归属人ID
                    'belong_user':obj.belong_user.username, # 归属人名字
                    'befor_add_fans':obj.befor_add_fans,    # 加粉前 粉丝数量
                    'after_add_fans':obj.after_add_fans,    # 加分后 粉丝数量
                    'add_fans_num':obj.add_fans_num,        # 添加的粉丝数量
                    'xiongzhanghaoID':obj.xiongzhanghaoID,  # 熊掌号ID
                    'search_keyword':obj.search_keyword,    # 熊掌号搜索关键词
                    'status':obj.get_status_display(),
                    'status_id':obj.status,
                    'create_date':obj.create_date.strftime('%Y-%m-%d'),
                    'errorText':obj.errorText ,              # 错误日志

                })
            #  查询成功 返回200 状态码
            response.code = 200
            response.msg = '查询成功'
            response.data = {
                'ret_data': ret_data,
                'status':models.xzh_add_fans.status_choices,
                'count':count
            }
    else:
        response.code = 301
        response.data = json.loads(forms_obj.errors.as_json())
    return JsonResponse(response.__dict__)


#  增删改
#  csrf  token验证
@csrf_exempt
@account.is_token(models.xzh_userprofile)
def fans_oper(request, oper_type, o_id):
    response = Response.ResponseObj()
    user_id = request.GET.get('user_id')
    user_objs = models.xzh_userprofile.objects.filter(id=user_id)
    if request.method == "POST":
        form_data = {
            'o_id':o_id,
            'oper_user_id': request.GET.get('user_id'),
            'belong_user_id': request.POST.get('belong_user_id'),
            'add_fans_num': request.POST.get('add_fans_num'),
            'xiongzhanghaoID': request.POST.get('xiongzhanghaoID'),
            'search_keyword': request.POST.get('search_keyword'),
        }
        if oper_type == "add":
            print('form_data----->',form_data)
            #  创建 form验证 实例（参数默认转成字典）

            forms_obj = AddForm(form_data)
            if forms_obj.is_valid():
                models.xzh_add_fans.objects.create(**forms_obj.cleaned_data)
                response.code = 200
                response.msg = "添加成功"
            else:
                print("验证不通过")
                # print(forms_obj.errors)
                response.code = 301
                # print(forms_obj.errors.as_json())
                response.msg = json.loads(forms_obj.errors.as_json())

        elif oper_type == "update":
            # 获取需要修改的信息
            forms_obj = UpdateForm(form_data)
            if forms_obj.is_valid():
                print("验证通过")
                formObjs = forms_obj.cleaned_data
                models.xzh_add_fans.objects.filter(id=o_id).update(
                    belong_user_id=formObjs.get('belong_user_id'),
                    add_fans_num=formObjs.get('add_fans_num'),
                    xiongzhanghaoID=formObjs.get('xiongzhanghaoID'),
                    search_keyword=formObjs.get('search_keyword'),
                )
                response.code = 200
                response.msg = "修改成功"
            else:
                print("验证不通过")
                # print(forms_obj.errors)
                response.code = 301
                # print(forms_obj.errors.as_json())
                #  字符串转换 json 字符串
                response.msg = json.loads(forms_obj.errors.as_json())

        elif oper_type == "delete":
            # 删除 ID
            if o_id == user_id:
                response.code = 301
                response.msg = '不能删除自己'
            else:
                objs = models.xzh_add_fans.objects.filter(id=o_id)
                if objs:
                    obj = objs[0]
                    obj.delete()
                    response.code = 200
                    response.msg = "删除成功"
                else:
                    response.code = 302
                    response.msg = '删除ID不存在'
            response.data = {}

    else:
        if oper_type == 'update_status':
            print('====')
            objs = models.xzh_add_fans.objects.filter(id=o_id)
            if objs:
                obj = objs[0]
                if obj.befor_add_fans:
                    status = 2
                else:
                    status = 1
                objs.update(status=status)
                response.code = 200
                response.msg = '修改成功'
            else:
                response.code = 301
                response.msg = '无修改ID'

        elif oper_type == 'exportExcle':
            print('=======================================================')
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="客户名称:")
            ws.cell(row=1, column=6, value="生成报表时间:")
            ws.cell(row=2, column=6, value="数据总数:")
            ws.cell(row=3, column=6, value="应加粉总数量:")
            ws.cell(row=4, column=6, value="实加粉总数量:")

            ws.cell(row=4, column=1, value="熊掌号搜索关键词")
            ws.cell(row=4, column=2, value="加粉前")
            ws.cell(row=4, column=3, value="加粉数")
            ws.cell(row=4, column=4, value="加粉后")
            ws.cell(row=4, column=5, value="加粉时间")


            # print('设置列宽')
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 25
            ws.column_dimensions['F'].width = 25
            ws.column_dimensions['G'].width = 25

            ws['A1'].alignment = Alignment(horizontal='right', vertical='center')
            ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['C1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['A2'].alignment = Alignment(horizontal='right', vertical='center')
            ws['C2'].alignment = Alignment(horizontal='right', vertical='center')
            ws['E1'].alignment = Alignment(horizontal='right', vertical='center')

            row = 5
            addObjs = models.xzh_add_fans.objects.filter(belong_user_id=o_id).order_by('-create_date')
            add_Objs = models.xzh_add_fans.objects.filter(belong_user_id=o_id).order_by('create_date')
            addObjsCount = addObjs.count()
            if addObjs:
                after_add_fans = addObjs[0].after_add_fans  #最后一次 加粉 数量
                befor_add_fans = add_Objs[0].befor_add_fans # 加粉前数量
                print(after_add_fans, befor_add_fans)
                now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ws.cell(row=1, column=2, value="{}".format(addObjs[0].belong_user.username))
                ws.cell(row=1, column=7, value="{}".format(now))
                ws.cell(row=2, column=7, value="{}".format(addObjsCount))
                ws.cell(row=4, column=7, value="{}".format(after_add_fans-befor_add_fans))
                yingjiafen_num = 0
                for addObj in addObjs:
                    yingjiafen_num += addObj.add_fans_num
                    print('addObj.create_date===========================> ',addObj.create_date)
                    ws.cell(row=row, column=1, value="{}".format(addObj.search_keyword))
                    ws.cell(row=row, column=2, value="{}".format(addObj.befor_add_fans))
                    ws.cell(row=row, column=3, value="{}".format(addObj.add_fans_num))
                    ws.cell(row=row, column=4, value="{}".format(addObj.after_add_fans))
                    ws.cell(row=row, column=5, value="{}".format(addObj.create_date))
                    row += 1
                ws.cell(row=3, column=7, value="{}".format(yingjiafen_num))

                # # 合并单元格        开始行      结束行       用哪列          占用哪列
                ws.merge_cells(start_row=1, end_row=2, start_column=1, end_column=1)
                ws.merge_cells(start_row=1, end_row=2, start_column=2, end_column=4)
                timestamp = str(int(time.time())) + str(time.time()).split('.')[1][2:-1]
                xlsx_url = 'statics/fansExcle/{}.xlsx'.format(timestamp)
                wb.save(xlsx_url)
                # wb.save('./5.xlsx')
                return_xlsx_path = 'http://xiongzhanghao.zhugeyingxiao.com:8003/' + xlsx_url
                response.code = 200
                response.msg = '生成完成'
                response.data = return_xlsx_path
            else:
                response.code = 301
                response.msg = '该用户无加粉数据'

        # 暂停加粉
        elif oper_type == 'pausePowder':
            print('===============================')
            oper_user = user_objs[0].username
            fans_objs = models.xzh_add_fans.objects.filter(id=o_id)
            if fans_objs:
                if int(fans_objs[0].status) == 2:
                    fans_objs.update(
                        status=5,
                        errorText='该任务被：{}暂停'.format(oper_user)
                    )
                    response.code = 200
                    response.msg = '暂停成功'
                elif int(fans_objs[0].status) == 5:
                    fans_objs.update(
                        status=2,
                        errorText=''
                    )
                    response.code = 200
                    response.msg = '开始成功'
                else:
                    response.code = 301
                    response.msg = '该状态不能暂停'
            else:
                response.code = 301
                response.msg = '无该任务'
        else:
            response.code = 402
            response.msg = "请求异常"

    return JsonResponse(response.__dict__)














