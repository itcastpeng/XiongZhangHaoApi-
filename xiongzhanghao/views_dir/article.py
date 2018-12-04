from xiongzhanghao import models
from xiongzhanghao.publicFunc import Response
from xiongzhanghao.publicFunc import account
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from xiongzhanghao.publicFunc.condition_com import conditionCom
from xiongzhanghao.forms.article import AddForm, UpdateForm, SelectForm
import json, datetime, requests, os
from django.db.models import Q
from backend.articlePublish import DeDe
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


# from xiongzhanghao.views_dir.user import objLogin


# cerf  token验证 用户展示模块
@csrf_exempt
@account.is_token(models.xzh_userprofile)
def article(request):
    response = Response.ResponseObj()
    if request.method == "GET":
        forms_obj = SelectForm(request.GET)
        if forms_obj.is_valid():
            current_page = forms_obj.cleaned_data['current_page']
            length = forms_obj.cleaned_data['length']
            print('forms_obj.cleaned_data -->', forms_obj.cleaned_data)
            order = request.GET.get('order', '-create_date')
            user_id = request.GET.get('user_id')
            start_time = request.GET.get('start_time')
            field_dict = {
                'id': '',
                'title': '__contains',
                'create_date': '',
                'summary': '__contains',
                'content': '__contains',
                'article_status': '',
                'belongToUser_id': '',
            }
            q = conditionCom(request, field_dict)
            print('start_time============>',start_time)
            if start_time:
                start_time = json.loads(start_time)
                stop_time = start_time[1]
                start_time = start_time[0]
                q.add(Q(create_date___gte=start_time) & Q(create_date__lte=stop_time), Q.AND)

            print('q -->', q)
            objs = models.xzh_article.objects.select_related('user', 'belongToUser').filter(q).order_by(order)
            count = objs.count()

            if length != 0:
                start_line = (current_page - 1) * length
                stop_line = start_line + length
                objs = objs[start_line: stop_line]

            # 返回的数据
            ret_data = []

            for obj in objs:
                # print('obj.id--------------> ',obj.id)
                #  将查询出来的数据 加入列表
                column = eval(obj.column_id) if obj.column_id else {}
                # print('column============> ', column)
                back_url = obj.back_url if obj.back_url else ''
                articlePicName = ''
                if obj.articlePicName:
                    articlePicName = obj.articlePicName
                send_time = obj.send_time.strftime('%Y-%m-%d %H:%M:%S') if obj.send_time else ''
                belongToUser = ''
                belongToUser_id = ''
                if obj.belongToUser:
                    belongToUser_id = obj.belongToUser_id
                    belongToUser = obj.belongToUser.username
                ret_data.append({
                    'id': obj.id,
                    'title':obj.title,
                    'summary':obj.summary,
                    'content':obj.content,
                    'column_id':column.get('Id'),
                    'column_name':column.get('name'),
                    'create_date':obj.create_date.strftime('%Y-%m-%d %H:%M:%S'),
                    'user_id':obj.user.id,
                    'user_name':obj.user.username,
                    'belongToUser_id':belongToUser_id,
                    'belongToUser_name': belongToUser,
                    'article_status': obj.get_article_status_display(),
                    'note_content':obj.note_content,
                    'back_url':back_url,
                    'send_time':send_time,
                    'is_audit':obj.is_audit,
                    'article_status_id':obj.article_status,
                    'is_delete':obj.is_delete,
                    'manualRelease':obj.manualRelease,
                    'articlePicName':articlePicName,
                })
            #  查询成功 返回200 状态码
            response.code = 200
            response.msg = '查询成功'
            response.data = {
                'ret_data': ret_data,
                'data_count': count,
                'article_status':models.xzh_article.article_status_choices,
            }
        else:
            response.code = 402
            response.msg = "请求异常"
            response.data = json.loads(forms_obj.errors.as_json())
    return JsonResponse(response.__dict__)


#  增删改
#  csrf  token验证
@csrf_exempt
@account.is_token(models.xzh_userprofile)
def article_oper(request, oper_type, o_id):
    response = Response.ResponseObj()
    if request.method == "POST":
        back_url = request.POST.get('back_url')  # 如果手动发布 回链必填
        manualRelease = request.POST.get('manualRelease')
        user_id = request.GET.get('user_id')
        belongToUser_id = request.POST.get('belongToUser_id')
        articlePicName = request.POST.get('articlePicName')    # 文章缩略图
        print('articlePicName=====================> ',articlePicName)
        if articlePicName:
            if ('http://www.zjnbsznfk120.com' or 'http://www.zjsznnk.com') in articlePicName:
                pass
            else:
                response.code = 301
                response.msg = '缩略图异常'
                return JsonResponse(response.__dict__)
        form_data = {
            'user_id': user_id,
            'title': request.POST.get('title'),
            'summary': request.POST.get('summary'),
            'content': request.POST.get('content'),
            'column_id': request.POST.get('column_id'),
            'create_date': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'belongToUser_id': belongToUser_id,
            'send_time': request.POST.get('send_time'),
            'manualRelease': manualRelease,
            'articlePicName':articlePicName
        }

        # print('form_data===============> ',form_data)
        if oper_type == "add":
            #  创建 form验证 实例（参数默认转成字典）
            forms_obj = AddForm(form_data)
            if forms_obj.is_valid():
                print("验证通过")
                # print("forms_obj.data.get('column_id')========> ",forms_obj.cleaned_data.get('articlePicName'))
                obj = models.xzh_article.objects.create(**forms_obj.cleaned_data)
                if manualRelease == 'true':
                    print('手动添加===========手动添加===========================手动添加')
                    if not back_url:
                        response.msg = '回链不能为空'
                        response.code = 301
                        return JsonResponse(response.__dict__)

                    articleObjs = models.xzh_article.objects.filter(id=obj.id)
                    article_status = 4
                    if int(articleObjs[0].belongToUser.userType) == 2:  # 特殊用户
                        article_status = 6
                    articleObjs.update(
                        article_status=article_status,
                        back_url=back_url,
                        is_audit=True,
                        articlePicName=articlePicName,  # 缩略图
                    )
                response.code = 200
                response.msg = "添加成功"
            else:
                print("验证不通过")
                response.code = 301
                response.msg = json.loads(forms_obj.errors.as_json())

        elif oper_type == "update":
            # 获取需要修改的信息
            forms_obj = UpdateForm(form_data)
            if forms_obj.is_valid():
                print("验证通过")
                #  查询数据库  用户id
                objs = models.xzh_article.objects.filter(
                    id=o_id
                )
                #  更新 数据
                if objs:
                    print('objs[0].article_status===============> ',objs[0].article_status)
                    if objs[0].article_status != 2:
                        objForm = forms_obj.cleaned_data
                        send_time = objForm.get('send_time')
                        objs.update(
                            user_id =objForm.get('user_id'),
                            title = objForm.get('title'),
                            summary = objForm.get('summary'),
                            content = objForm.get('content'),
                            belongToUser_id = objForm.get('belongToUser_id'),
                            column_id = objForm.get('column_id'),
                            back_url=back_url,
                            articlePicName=articlePicName,  # 缩略图
                        )
                        if send_time:
                            objs.update(send_time=send_time)

                        response.code = 200
                        response.msg = "修改成功"
                    else:
                        response.code = 301
                        response.msg = '发布成功, 不可修改'
                else:
                    response.code = 303
                    response.msg = json.loads(forms_obj.errors.as_json())

            else:
                print("验证不通过")
                # print(forms_obj.errors)
                response.code = 301
                # print(forms_obj.errors.as_json())
                #  字符串转换 json 字符串
                response.msg = json.loads(forms_obj.errors.as_json())

        elif oper_type == "delete":
            # 删除 ID
            company_id = request.GET.get('company_id')
            objs = models.xzh_article.objects.filter(id=o_id)
            if objs:
                objs.delete()
                response.code = 200
                response.msg = "删除成功"
            else:
                response.code = 302
                response.msg = '删除ID不存在'

        # 重新发布文章
        elif oper_type == 'redistribution':
            objs = models.xzh_article.objects.filter(id=o_id)
            if objs[0].article_status != 2:
                objs.update(article_status=1,note_content='')
            response.code = 200
            response.msg = '重新发布成功'

        # 下载文章统计报表
        elif oper_type == 'exportExcel':
            o_id = request.POST.get('o_id')
            date_time = request.POST.get('date_time')
            start_time = request.POST.get('start_time')
            stop_time = request.POST.get('stop_time')
            if o_id:
                now_date = datetime.datetime.now()
                objs = models.xzh_article.objects.filter(belongToUser_id=o_id).order_by('create_date')
                obj = objs[0]
                q = Q()
                start = obj.create_date
                if date_time and int(date_time):
                    stop_now = datetime.datetime.now().strftime('%Y-%m-%d 23:59:59')
                    start_now = datetime.datetime.now().strftime('%Y-%m-%d 00:00:00')
                    stop = stop_now
                    if int(date_time) == 1:
                        start = start_now
                    elif int(date_time) == 7:
                        start = (now_date - datetime.timedelta(days=7)).strftime('%Y-%m-%d %H:%M:%S')
                    elif int(date_time) == 30:
                        start = (now_date - datetime.timedelta(days=30)).strftime('%Y-%m-%d %H:%M:%S')
                    q.add(Q(belongToUser_id=o_id) & Q(create_date__lte=stop) & Q(create_date__gte=start), Q.AND)
                elif start_time and stop_time:
                    q.add(Q(belongToUser_id=o_id) & Q(create_date__lte=stop_time) & Q(create_date__gte=start_time), Q.AND)
                else:
                    response.code = 301
                    response.msg = '下载报表异常！'
                print('q-----------> ', q)
                wb = Workbook()
                ws = wb.active
                ws.title = '关键词覆盖查询'
                ws.cell(row=1, column=1, value="用户:")
                ws.cell(row=2, column=1, value="报表生成时间:")
                ws.cell(row=1, column=3, value="数据总数:")
                ws.cell(row=4, column=1, value="---创建时间---")
                ws.cell(row=4, column=2, value="---文章标题---")
                ws.cell(row=4, column=3, value="---所选栏目---")
                ws.cell(row=4, column=4, value="---回链地址---")

                # # 合并单元格        开始行      结束行       用哪列          占用哪列
                ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

                # print('设置列宽')
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 30
                ws.column_dimensions['C'].width = 30
                ws.column_dimensions['D'].width = 50

                # 文本居中
                ws['A1'].alignment = Alignment(horizontal='right', vertical='center')
                ws['A2'].alignment = Alignment(horizontal='right', vertical='center')
                ws['C1'].alignment = Alignment(horizontal='right', vertical='center')

                objs = models.xzh_article.objects.filter(q)
                row = 5
                now_date = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                if objs:
                    username = objs[0].belongToUser.username
                    ws.cell(row=1, column=2, value="{now_date}".format(now_date=username))
                    ws.cell(row=2, column=2, value="{now_date}".format(now_date=now_date))
                    ws.cell(row=1, column=4, value="{now_date}".format(now_date=objs.count()))
                    for obj in objs:
                        column = ''
                        if obj.column_id:
                            column = eval(obj.column_id)['name']
                        ws.cell(row=row, column=1, value="{create_time}".format(create_time=obj.create_date))
                        ws.cell(row=row, column=2, value="{title}".format(title=obj.title))
                        ws.cell(row=row, column=3, value="{column_id}".format(column_id=column))
                        ws.cell(row=row, column=4, value="{back_url}".format(back_url=obj.back_url))
                        row += 1

                    path_name = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
                    name = username + '---' + datetime.datetime.now().strftime('%Y-%m-%d')
                    path = os.path.join(path_name, 'statics', 'wenzhangbaobiao', name + '.xlsx')
                    wb.save(path)
                    # return_path = 'http://192.168.10.207:8003/statics/wenzhangbaobiao/' + name + '.xlsx'
                    return_path = 'http://xiongzhanghao.zhugeyingxiao.com:8003/statics/wenzhangbaobiao/' + name + '.xlsx'
                    print('return_path---> ', return_path)
                    response.code = 200
                    response.msg = '生成成功'
                    response.data = return_path
                else:
                    response.code = 301
                    response.msg = '该用户没有发布任何文章！'
            else:
                response.code = 301
                response.msg = '无该用户！'

        # celery 调用异步统计报表
        elif oper_type == 'celeryExportExcel':
            start = request.POST.get('start')
            stop = request.POST.get('stop')
            o_id = request.POST.get('o_id')
            if o_id:
                wb = Workbook()
                ws = wb.active
                ws.title = '关键词覆盖查询'
                ws.cell(row=1, column=1, value="用户:")
                ws.cell(row=2, column=1, value="报表生成时间:")
                ws.cell(row=1, column=3, value="数据总数:")
                ws.cell(row=4, column=1, value="---创建时间---")
                ws.cell(row=4, column=2, value="---文章标题---")
                ws.cell(row=4, column=3, value="---所选栏目---")
                ws.cell(row=4, column=4, value="---回链地址---")

                # # 合并单元格        开始行      结束行       用哪列          占用哪列
                ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=1)

                # print('设置列宽')
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 30
                ws.column_dimensions['C'].width = 30
                ws.column_dimensions['D'].width = 50

                # 文本居中
                ws['A1'].alignment = Alignment(horizontal='right', vertical='center')
                ws['A2'].alignment = Alignment(horizontal='right', vertical='center')
                ws['C1'].alignment = Alignment(horizontal='right', vertical='center')
                q = Q()
                q.add(Q(belongToUser_id=o_id) & Q(create_date__lte=stop) & Q(create_date__gte=start), Q.AND)
                print('q-----------> ', q)
                objs = models.xzh_article.objects.filter(q)
                row = 5
                now_date = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                if objs:
                    username = objs[0].belongToUser.username
                    ws.cell(row=1, column=2, value="{now_date}".format(now_date=username))
                    ws.cell(row=2, column=2, value="{now_date}".format(now_date=now_date))
                    ws.cell(row=1, column=4, value="{now_date}".format(now_date=objs.count()))
                    for obj in objs:
                        column = ''
                        if obj.column_id:
                            column = eval(obj.column_id)['name']
                        ws.cell(row=row, column=1, value="{create_time}".format(create_time=obj.create_date))
                        ws.cell(row=row, column=2, value="{title}".format(title=obj.title))
                        ws.cell(row=row, column=3, value="{column_id}".format(column_id=column))
                        ws.cell(row=row, column=4, value="{back_url}".format(back_url=obj.back_url))
                        row += 1

                    path_name = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
                    # name = username
                    name = username + '---' + datetime.datetime.now().strftime('%Y-%m-%d')
                    path = os.path.join(path_name, 'statics', 'wenzhangbaobiao', name + '.xlsx')
                    print('path_name============> ',path)
                    wb.save(path)
                    response.code = 200


                else:
                    response.code = 301
                    response.msg = '该用户没有发布任何文章！'

        else:
            response.code = 402
            response.msg = '请求异常'


    else:
        if oper_type == 'thumbnail':   # 查询缩略图  妇科
            # ===========================================该代码 爬取缩略图使用=====================================
            # response.code = 200
            # response.msg = '查询成功'
            # cookie = {'DedeLoginTime': '1543226988', 'DedeUserID__ckMd5': 'eeead0cc3feb9247',
            #           'PHPSESSID': 'aocgl1iqc56p18mm22epgglks4', 'DedeUserID': '1',
            #           'DedeLoginTime__ckMd5': '36b39a5c7cf5c4d6'}
            # pwd = 'tgb123qaz'
            # userid = 'admin'
            # domain = 'http://www.zjsznnk.com/'
            # home_path = 's_z_n_yy'
            #
            # dede = DeDe(domain, home_path, userid, pwd, cookie)
            # dede.login()
            # data_list = dede.suoluetu()
            # for i in data_list:
            #     p = '/uploads' + i
            #     models.xzh_suoluetu.objects.create(
            #         man=p
            #     )
            # ============================================================================================
            img_list = []
            objs = models.xzh_suoluetu.objects.filter(man__isnull=True)
            for obj in objs:
                suoluetu = str('http://www.zjnbsznfk120.com' + obj.woman)
                if suoluetu not in img_list:
                    img_list.append(suoluetu)
            response.code = 200
            response.msg = '查询成功'
            response.data = img_list
        elif oper_type == 'thumbnailMan':  # 男科
            img_list = []
            objs = models.xzh_suoluetu.objects.filter(man__isnull=False)
            for obj in objs:
                suoluetu = str('http://www.zjsznnk.com' + obj.man)
                # print('suoluetu-------> ',suoluetu)
                if suoluetu not in img_list:
                    img_list.append(suoluetu)
            response.code = 200
            response.msg = '查询成功'
            response.data = img_list
        else:
            response.code = 402
            response.msg = "请求异常"
    return JsonResponse(response.__dict__)

