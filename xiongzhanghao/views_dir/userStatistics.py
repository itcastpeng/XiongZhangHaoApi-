from xiongzhanghao import models
from xiongzhanghao.publicFunc import Response
from xiongzhanghao.publicFunc import account
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from xiongzhanghao.publicFunc.condition_com import conditionCom
from xiongzhanghao.forms.add_fans import SelectForm
import json, requests, datetime, time
from openpyxl import Workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Font, Alignment
from django.db.models import Q



# cerf  token验证 用户展示模块
@csrf_exempt
@account.is_token(models.xzh_userprofile)
def userStatistics(request):
    response = Response.ResponseObj()
    forms_obj = SelectForm(request.GET)
    if forms_obj.is_valid():
        current_page = forms_obj.cleaned_data['current_page']
        length = forms_obj.cleaned_data['length']
        print('forms_obj.cleaned_data -->', forms_obj.cleaned_data)
        order = request.GET.get('order', '-create_date')
        days = request.GET.get('days', 7)
        field_dict = {
            'id': '',
            'username': '__contains',
            'create_date': '__contains',
            'belong_user_id': '',
        }
        q = conditionCom(request, field_dict)

        now = datetime.datetime.now()
        nowDate = now.strftime('%Y-%m-%d')
        if days:
            stop = nowDate
            if int(days) == 7:
                time_Y_M_D = (now - datetime.timedelta(days=6)).strftime('%Y-%m-%d')
            else:
                time_Y_M_D = (now - datetime.timedelta(days=30)).strftime('%Y-%m-%d')
            start = time_Y_M_D
            q.add(Q(create_date__lte=stop) & Q(create_date__gte=start), Q.AND)
        print('q -->', q)
        objs = models.user_statistics.objects.select_related('belong_user').filter(q).order_by(order)
        count = objs.count()

        if length != 0:
            start_line = (current_page - 1) * length
            stop_line = start_line + length
            objs = objs[start_line: stop_line]

        # 返回的数据
        ret_data = []

        for obj in objs:
            #  将查询出来的数据 加入列表
            ret_data.append({
                'id': obj.id,
                'belong_user_id':obj.belong_user_id,    # 归属人ID
                'belong_user':obj.belong_user.username, # 归属人名字
                'public_num':obj.public_num,            # 发布数量
                'fans_num':obj.fans_num,                # 粉丝数量
                'zhishu':obj.zhishu,                    # 指数
                'zhanxianliang':obj.zhanxianliang,      # 展现量
                'dianjiliang':obj.dianjiliang,          # 点击量
                'baidu_shoulu':obj.baidu_shoulu,        # 百度收录数量
                # 'baidu_shoulu_url':obj.baidu_shoulu_url,  # 百度收录链接
                'index_show':obj.index_show,            # 熊掌号主页展示条数 （主页收录）
                # 'index_show_url':obj.index_show_url,  # 熊掌号主页展示url
                'admin_shoulu':obj.admin_shoulu,        # 熊掌号后台收录条数
                # 'admin_shoulu_url':obj.admin_shoulu_url,  # 熊掌号后台收录url
                'create_date':obj.create_date.strftime('%Y-%m-%d'), # 创建时间
            })

        userObj = models.user_statistics.objects.values('belong_user_id', 'belong_user__username').distinct()  # 查询所有用户
        userList = []
        for i in userObj:
            userList.append({
                'id':i.get('belong_user_id'),
                'name':i.get('belong_user__username')
            })
        #  查询成功 返回200 状态码
        response.code = 200
        response.msg = '查询成功'
        response.data = {
            'ret_data': ret_data,
            'count':count,
            'userList':userList
        }

    else:
        response.code = 301
        response.data = json.loads(forms_obj.errors.as_json())

    return JsonResponse(response.__dict__)



#  增删改
#  csrf  token验证
@csrf_exempt
@account.is_token(models.xzh_userprofile)
def userStatistics_oper(request, oper_type, o_id):
    response = Response.ResponseObj()
    user_id = request.GET.get('user_id')
    if request.method == 'POST':

        # 生成报表
        if oper_type == 'userStatisticsExcle':
            days = request.GET.get('days', 7)
            now = datetime.datetime.now()
            nowDate = now.strftime('%Y-%m-%d')
            stop = nowDate
            if int(days) == 7:
                time_Y_M_D = (now - datetime.timedelta(days=6)).strftime('%Y-%m-%d')
            else:
                time_Y_M_D = (now - datetime.timedelta(days=30)).strftime('%Y-%m-%d')
            start = time_Y_M_D

            q = Q()
            if o_id:
                q.add(Q(belong_user_id=o_id), Q.AND)

            q.add(Q(create_date__lte=stop) & Q(create_date__gte=start), Q.AND)
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="客户名称:").font = Font('宋体', size=11, b=True)

            ws.cell(row=3, column=1, value="查询时间").font = Font('宋体', size=11, b=True)
            ws.cell(row=3, column=2, value="熊掌号指数").font = Font('宋体', size=11, b=True)
            ws.cell(row=3, column=3, value="粉丝数量").font = Font('宋体', size=11, b=True)
            ws.cell(row=3, column=4, value="发布数量").font = Font('宋体', size=11, b=True)
            ws.cell(row=3, column=5, value="收录数量").font = Font('宋体', size=11, b=True)
            ws.cell(row=3, column=6, value="主页展示").font = Font('宋体', size=11, b=True)
            ws.cell(row=3, column=7, value="点击量").font = Font('宋体', size=11, b=True)
            ws.cell(row=3, column=8, value="展现量").font = Font('宋体', size=11, b=True)

            # print('设置列宽')
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 15

            ws['A1'].alignment = Alignment(horizontal='right', vertical='center')
            ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['C1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
            ws['B3'].alignment = Alignment(horizontal='center', vertical='center')
            ws['C3'].alignment = Alignment(horizontal='center', vertical='center')
            ws['D3'].alignment = Alignment(horizontal='center', vertical='center')
            ws['E3'].alignment = Alignment(horizontal='center', vertical='center')
            ws['F3'].alignment = Alignment(horizontal='center', vertical='center')
            ws['G3'].alignment = Alignment(horizontal='center', vertical='center')
            ws['H3'].alignment = Alignment(horizontal='center', vertical='center')


            row = 4
            objs = models.user_statistics.objects.filter(q).order_by('-create_date')
            objCount = objs.count()
            obj = objs[0]
            now = datetime.datetime.now().strftime('%Y-%m-%d')
            ws.cell(row=1, column=2, value="{}".format(obj.belong_user.username)).font = Font('宋体', size=11, b=True)

            for obj in objs:
                zhanxianliang = obj.zhanxianliang
                dianjiliang = obj.dianjiliang
                if obj.create_date.strftime('%Y-%m-%d') == now:
                    zhanxianliang = '未出'
                    dianjiliang = '未出'
                if obj.admin_shoulu > obj.baidu_shoulu:
                    shoulu = obj.admin_shoulu
                else:
                    shoulu = obj.baidu_shoulu
                if shoulu < obj.index_show:
                    shoulu = obj.index_show
                ws['A' + str(row)].alignment = Alignment(horizontal='center', vertical='center')
                ws['B' + str(row)].alignment = Alignment(horizontal='center', vertical='center')
                ws['C' + str(row)].alignment = Alignment(horizontal='center', vertical='center')
                ws['D' + str(row)].alignment = Alignment(horizontal='center', vertical='center')
                ws['E' + str(row)].alignment = Alignment(horizontal='center', vertical='center')
                ws['F' + str(row)].alignment = Alignment(horizontal='center', vertical='center')
                ws['G' + str(row)].alignment = Alignment(horizontal='center', vertical='center')
                ws['H' + str(row)].alignment = Alignment(horizontal='center', vertical='center')

                ws.cell(row=row, column=1, value="{}".format(obj.create_date)).font = Font('宋体', size=10, b=True)
                ws.cell(row=row, column=2, value="{}".format(obj.zhishu)).font = Font('宋体', size=10, b=True)
                ws.cell(row=row, column=3, value="{}".format(obj.fans_num)).font = Font('宋体', size=10, b=True)
                ws.cell(row=row, column=4, value="{}".format(obj.public_num)).font = Font('宋体', size=10, b=True)
                ws.cell(row=row, column=5, value="{}".format(shoulu)).font = Font('宋体', size=10, b=True)
                ws.cell(row=row, column=6, value="{}".format(obj.index_show)).font = Font('宋体', size=10, b=True)
                ws.cell(row=row, column=7, value="{}".format(dianjiliang)).font = Font('宋体', size=10, b=True)
                ws.cell(row=row, column=8, value="{}".format(zhanxianliang)).font = Font('宋体', size=10, b=True)
                row += 1

            # # 合并单元格        开始行      结束行       用哪列          占用哪列
            ws.merge_cells(start_row=1, end_row=2, start_column=1, end_column=1)
            ws.merge_cells(start_row=1, end_row=2, start_column=2, end_column=8)



            timestamp = str(int(time.time())) + str(time.time()).split('.')[1][2:-1]
            xlsx_url = 'statics/fansExcle/{}.xlsx'.format(timestamp)
            wb.save(xlsx_url)
            # wb.save('./5.xlsx')
            return_xlsx_path = 'http://xiongzhanghao.zhugeyingxiao.com:8003/' + xlsx_url
            response.code = 200
            response.msg = '生成完成'
            response.data = return_xlsx_path

    else:
        response.code = 402
        response.msg = '请求错误'
    return JsonResponse(response.__dict__)














