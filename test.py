
import requests, json
from bs4 import BeautifulSoup





from backend.lianzhongDama import LianZhongDama



# data = {
#     'dosubmit':'',
#     'username':'Ymexiongzhanghao',
#     'password':'Yme@evercare20181126',
#     'code':''
#
# }




# url = 'http://m.evercarebj.com/index.php?m=admin&c=index&a=login%dosubmit=1'
# ret2 = requests.post(url, data=data)
# print(ret2.text)



class PcV9(object):
    def __init__(self, userid, pwd, cookies=None):
        self.requests_obj = requests.session()
        self.userid = userid
        self.pwd = pwd
        self.cookies = cookies


    def login(self):
        url = 'http://m.5iyme.com/api.php?op=checkcode&code_len=4&font_size=20&width=130&height=50&font_color=&background='  # 获取验证码
        login_url = 'http://m.5iyme.com/index.php?m=admin&c=index&a=login&dosubmit=1'  # 登录
        if self.cookies:
            pc_hash = self.is_login(login_url)
            return self.cookies, pc_hash
        while True:
            try:
                yzmRet = self.requests_obj.get(url)
                with open('yzm.jpg', 'wb') as f:
                    f.write(yzmRet.content)
                obj = LianZhongDama()
                yzm = obj.get_yzm('yzm.jpg')
                print('yzm -->', yzm)
                break
            except AttributeError:
                break
        yzm = ''
        post_data = {
            'dosubmit': '',
            'username': self.userid,
            'password': self.pwd,
            'code': '{}'.format(yzm)
        }
        ret = self.requests_obj.post(login_url, data=post_data)
        soup = BeautifulSoup(ret.text, 'lxml')
        pc_hash = soup.find('a', text='如果您的浏览器没有自动跳转，请点击这里')
        pc_hash = pc_hash.attrs.get('href').split('pc_hash=')[-1]
        print('ret.text----------->',ret.text)
        # print('self.requests_obj.cookies---------> ',self.requests_obj.cookies)
        # print('self.requests_obj.cookies---------> ',self.requests_obj.headers)
        cookies = requests.utils.dict_from_cookiejar(self.requests_obj.cookies)
        self.cookies = cookies
        print('cookies============> ',cookies)
        return cookies, pc_hash

    # 判断是否登录
    def is_login(self, login_url):
        post_data = {
            'dosubmit': '',
            'username': self.userid,
            'password': self.pwd,
        }
        # print('判断是否登录========》 ', login_url, self.cookies)
        ret = self.requests_obj.post(login_url, cookies=self.cookies, data=post_data)
        if '登录成功' in ret.text:
            # print(ret.text)
            soup = BeautifulSoup(ret.text, 'lxml')
            pc_hash = soup.find('a', text='如果您的浏览器没有自动跳转，请点击这里')
            pc_hash = pc_hash.attrs.get('href').split('pc_hash=')[-1]
            return pc_hash
        if "返回上一页" in ret.text:
            self.cookies = ''    # cookie清除掉
            self.login()

    # 获取栏目信息
    def getClassInfo(self, pc_hash):
        url = 'http://m.evercarebj.com/index.php?m=content&c=content&a=public_categorys&type=add&menuid=822&pc_hash={}'.format(pc_hash)
        ret = self.requests_obj.get(url, cookies=self.cookies)
        soup = BeautifulSoup(ret.text, 'lxml')
        category_tree = soup.find('ul', id='category_tree')
        li_tags = category_tree.find_all('li')
        retData = []
        for li_tag in li_tags:
            a_tag = li_tag.find('span', class_='')
            class_id = a_tag.find('a').attrs.get('href').split('&')[-1].split('=')[-1]
            class_name = a_tag.get_text()
            print('class_id-------class_name---> ',class_id, class_name)

            data_list = [class_id, class_name]

            if data_list not in retData:
                retData.append(
                    data_list
                )
        return retData

    # # 判断标题是否可用
    def article_test_title(self, catid, title):
        print("判断标题是否可用")
        url = 'http://m.evercarebj.com/index.php?m=content&c=content&a=public_check_title&catid={}&sid=2.037687345459017&data={}'.format(catid, title)
        ret = self.requests_obj.get(url, cookies=self.cookies)
        print(ret.text)
        if ret.text == '0':
            return True   # 不重复
        return False

    # # 发布文章
    def sendArticle(self, data, title, pc_hash):
        catid = data.get('info[catid]')
        if self.article_test_title(catid, title):
            print('增加文章')
            data['pc_hash'] = pc_hash
            url = 'http://m.5iyme.com/index.php?m=content&c=content&a=add'
            print('发布url==================> ',url)
            ret = self.requests_obj.post(url, data=data, cookies=self.cookies)
            print('ret.text--> ', ret.text)
            if '数据添加成功' in ret.text:
                'http://m.5iyme.com/index.php?m=admin&c=index&pc_hash=G0Snyn'
                url = 'http://m.5iyme.com/index.php?m=content&c=content&a=init&menuid=822&catid={}&pc_hash={}'.format(
                    data.get('info[catid]'), pc_hash)
                print('url--------------------> ',url)
                ret = self.requests_obj.get(url, cookies=self.cookies)
                title = data.get('info[title]')
                print('=============title================title=================> ',title)
                if title in ret.text:
                    print('---------------------------发布成功-------------------------')
                    soup = BeautifulSoup(ret.text, 'lxml')
                    tr_all = soup.find_all('tr')
                    aid = 0
                    huilian = ''
                    for tr in tr_all:
                        if title.strip() in tr.get_text():
                            td_all = tr.find_all('td')
                            aid = td_all[2].get_text()
                            huilian = td_all[3].find('a').attrs.get('href')
                            break
                    return {
                        'huilian':huilian,
                        'aid':aid,
                        'code':200
                    }
                else:
                    print('标题未==查到===============================标题未查到==========================标题未查到')
                    return {
                        'huilian': '',
                        'code': 301
                    }
            else:
                print('===============添加文章失败=============')
                return {
                    'huilian': '',
                    'code': 500
                }
        else:
            print('’发布失败=========================发布失败===================发布失败 300')
            return {
                'huilian': '',
                'code': 300
                    }


    def deleteQuery(self, url, maxtime, data_list):
        print('url--------------------> ', url)
        ret = self.requests_obj.get(url, cookies=self.cookies)
        soup = BeautifulSoup(ret.text, 'lxml')
        tr_all = soup.find_all('tr')
        for tr in tr_all:
            if tr.find_all('td', align='center'):
                aid = tr.find_all('td')[2].get_text().strip()
                title = tr.find_all('td')[3].get_text().strip()
                date_time = tr.find_all('td')[6].get_text()
                if date_time < maxtime:
                    continue
                if aid and title:
                    data_list.append({
                        'aid':aid,
                        'title':title,
                        'date_time':date_time
                    })
                else:
                    continue
        next = soup.find('a', text='下一页')
        if next:
            next_page = soup.find('a', text='下一页')
            next_url = next_page.attrs.get('href')
            if next_url == url:
                return data_list
            print('next_url--> ',next_url)
            self.deleteQuery(next_url, maxtime, data_list)
        return data_list

import datetime
if __name__ == '__main__':
    catid = 115
    title = '测试标题'
    user_id = 'Ymexiongzhanghao'
    password = 'Yme@evercare20181126'
    cookie = {'PHPSESSID': 'q5981m9bglfllmc3bib4lvuti6',
              'Hhuvt_sys_lang': 'a5e0FMsBPHEjViLciHcdIze2UdMxPOkOwJq4edMDAAwEiw',
              'Hhuvt_admin_username': 'e242V5S-mKfYXHGRRUODNuUI42_b-YJ_XHZcoxfaiEzJ03n62WyVb9vuZ-d9',
              'Hhuvt_siteid': '9222HgxnLMBETl68Jcw2DQcV_cAY8T3ZF6KeueBl',
              'Hhuvt_admin_email': '0e467jj9jrpv_NDlXj-m_73TUGIWn3ab4iKirTA4qN-ElYcCtAZcBwx-w_Y',
              'Hhuvt_userid': '9b82AUqls4zcOtqf33LDiw-ZpeOSjZIx9gLE6NOFAA'}
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    data = {
        'info[thumb]':'',
        'info[relation]':'',
        'info[inputtime]':now,
        'info[islink]':0,
        'info[template]':'',
        'info[allow_comment]':1,
        'info[readpoint]':'',
        'info[paytype]':0,
        'info[catid]':catid,
        'info[title]':title,
        'style_color':'',
        'style_font_weight':'',
        'info[keywords]':'',
        'info[copyfrom]':'',
        'copyfrom_data':0,
        'info[description]':'测试摘要====',
        'info[content]':'测试内容----内容',
        'page_title_value':'',
        'add_introduce':1,
        'introcude_length':200,
        'auto_thumb':1,
        'auto_thumb_no':1,
        'info[paginationtype]':0,
        'info[maxcharperpage]':10000,
        'info[posids][]':-1,
        'info[groupids_view]':1,
        'info[voteid]':'',
        'dosubmit':'保存后自动关闭',
    }
    objs = PcV9(user_id, password, cookie)
    cookie, pc_hash = objs.login()
    # print('pc_hash========> ',cookie, pc_hash)
    # objs.getClassInfo()
    objs.sendArticle(data, title, pc_hash)
    # 判断是否删除
    # maxtime = '2018-11-20'

    # url = 'http://m.evercarebj.com/index.php?m=content&c=content&a=init&menuid=822&catid={}&pc_hash={}'.format(
    #     catid, pc_hash)
    # data_list = []
    # data_list = objs.deleteQuery(url, maxtime, data_list)
    # objs.getClassInfo(pc_hash)
    # print(data_list)
# import time
#
# from xiongzhanghao.publicFunc.account import str_encrypt
#
# token = 'a66b1a82b4ba3ca9d444322c8524e844'
# timestamp = str(int(time.time() * 1000))
# params = {
#     'user_id': 44,
#     'rand_str': str_encrypt(timestamp + token),
#     'timestamp': timestamp,
# }
# print(params)



from openpyxl.styles import Font, Alignment
from openpyxl import Workbook


wb = Workbook()
ws = wb.active
ws.title = '关键词覆盖查询'
ws.cell(row=1, column=1, value="用户:")
ws.cell(row=3, column=1, value="创建时间")
ws.cell(row=3, column=2, value="文章标题")
ws.cell(row=3, column=3, value="所选栏目")
ws.cell(row=3, column=4, value="回链地址")
# ft1 = Font(name='宋体', size=22)
# a1 = ws['A1']
# a1.font = ft1

# # 合并单元格        开始行      结束行       用哪列          占用哪列
ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=2)

# ws.merge_cells(start_row=2, end_row=5, start_column=5, end_column=5)

# print('设置列宽')
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 30
ws.column_dimensions['E'].width = 30
ws.column_dimensions['F'].width = 30
ws.column_dimensions['G'].width = 30


# # print('设置行高')
# ws.row_dimensions[1].height = 28

# # print('文本居中')
# ws['A1'].alignment = Alignment(horizontal='center', vertical='center')


# wb.save('./1.xlsx')


# token = 'a66b1a82b4ba3ca9d444322c8524e844'
# timestamp = str(int(time.time() * 1000))
# params = {
#     'user_id': 44,
#     'rand_str': str_encrypt(timestamp + token),
#     'timestamp': timestamp,
# }
#
# api_url = "http://xiongzhanghao.zhugeyingxiao.com:8003/api/select_keywords_cover"
# print(api_url)
# ret = requests.get(api_url, params=params, timeout=30)
# print(ret.text)
# result_data = json.loads(ret.text)
#
# if result_data["code"] == 200 and len(result_data["data"]) > 0:
#     print('=====================')



from xiongzhanghao.publicFunc.account import str_encrypt
import time
xiongzhanghao_url = 'https://xiongzhang.baidu.com/site/login'
# ret = requests.get(xiongzhanghao_url)
# encode_ret = ret.apparent_encoding
# # print('encode_ret===========', encode_ret)
# if encode_ret == 'GB2312':
#     ret.encoding = 'gbk'
# else:
#     ret.encoding = 'utf-8'
# print(ret.text)


# xiongzhanghao = 'https://xiongzhang.baidu.com/site/exponent?officeId=1610219546234420'


# ret = requests.get(xiongzhanghao_url)
# ret.encoding = 'utf8'
# print(ret.text)


# url = 'http://xiongzhanghao.zhugeyingxiao.com:8003/api/user_statistical/getTask?user_id=44&timestamp=1542788198850&rand_str=86b24054d91240d9559e369296af06cd'
# ret = requests.get(url, timeout=5)
# result_data = ret.json()
# print('result_data -->', result_data)
# if result_data['code'] == 200:
#     if result_data['data']['flag']:
#         print('=========================')




cookies = {
    'PSTM':'1531962644',
    'BIDUPSID':'34856101788F122E450B377F9896FF34',
    'BAIDUID':'C6A5A95F520664BCB3C91843B00D7D83:SL=0:NR=10:FG=1',
    'delPer':'0',
    'PHPSESSID':'1653161623663383339363464693435643934323367346368363668366137316',
    'BDUSS':'FMUENsZmUtSU9PejF5Qk1ZQUtTdUNxY350c0ZhZnQ3eVpQeDg2LWNOc1VaREZjQVFBQUFBJCQAAAAAAAAAAAEAAAAgH-aeAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABTXCVwU1wlcRF',
    's_id_ext1':'4064d5bca141bd59f77c5e59cbd574f3',
    'FEED_SIDS':'961300_1207_10',
    'BDORZ':'AE84CDB3A529C0F8A2B9DCDD1D18B695',
    'SE_LAUNCH':'5%3A25735821_0%3A25735821',
    'BDPASSGATE':'IlPT2AEptyoA_yiU4VOu3kIN8efRWv3A1hHJSFptQVStfCaWmhH3U0VlQDPHTnOZHpD6xpWro_-yoUbCQjpfjKMThgMAfztJbEfg-M725aTvL16dfLco_2z5V6g4trnG8ahgxwoT_eVEVFoKewPfogYIehS3ywxHafmI7X0fh2Ll_TCG17r6r7iDZ9sJLCPAPNu594rSjEdgXJmfA4H6TS0dmUsPKCIYzNzbdsExOevWrnW',
    'H_WISE_SIDS':'126893_125821_127696_127288_127237_127418_126427_123252_127486_126255_120172_123019_127529_118881_118861_118845_118823_118787_127181_107317_126995_126864_126145_127569_126797_117332_126794_117428_126837_126783_126442_127415_127029_126380_127597_127323_126684_126282_126942_126959_126773_126720_126091_125853_127536_124800_126437_110085_126973_123290_127699_127032_127225_100457',
    'PSINO':'2',
    'BDSVRBFE':'Go',
    '__bsi':'5989811770223376154_h2_2_R_R_30_0303_cca8_Y',
}


# # url = 'https://xiongzhang.baidu.com/account/exp/current?ajax=1&officeId=1610219546234420&bdstoken=89f3efad01ed08bc08cb74f3436a4764' # 熊掌号指数
# url = 'https://xiongzhang.baidu.com/account/exp/current?ajax=1&officeId=1610219546234420&bdstoken=721ac7674f9edbb80493fa9780bd8830' # 熊掌号指数
# ret = requests.get(url, cookies=cookies)
# print(ret.text)
# zhishu = ret.json().get('data').get('exp')
#
#
# url = 'https://ziyuan.baidu.com/xzh/analysis/stat?appid=1610219546234420&type=dispClickAll&start=&end=&period=week'
# ret = requests.get(url, cookies=cookies)
# ret.encoding = 'utf8'
# ret_list = ret.json().get('list')
# # for i in ret_list:
# #     print(i.get('fullDayClick'))   # 点击量
# #     print(i.get('fullDayDisplay'))   # 展现量
# #     print(i.get('date'))             # 时间
# #     print(i.get('pecent'))           # 点展比
# print(ret_list)

from ftplib import FTP
import logging.config

# 将日志写入文件
# logger = logging.getLogger(__name__)
# handler = logging.FileHandler("log.txt")
# handler.setLevel(logging.INFO)
# formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
# handler.setFormatter(formatter)
# logger.addHandler(handler)


# host = '61.188.39.201'
# port = 1987
# username = 'kmxzh'
# password = 'kmdyzXZH123'
# file = '1.txt'

# f = FTP()  # 实例化FTP对象
# f.connect(host, port)
# f.login(username, password)  # 登录

# pwd_path = f.pwd()
# print("FTP当前路径:", pwd_path)
#
#
# file_remote = '1.txt'
# file_local = 'D:\\test_data\\ftp_download.txt'
# bufsize = 1024
#
# fp = open(file_local, 'wb')
#
# f.retrbinary('RETR %s' % file_remote, fp.write, bufsize)
#
# fp.close()



# file_remote = 'ftp_upload.txt'
# file_local = 'D:\\test_data\\ftp_upload.txt'
# bufsize = 1024  # 设置缓冲器大小
# fp = open(file_local, 'rb')
# f.storbinary('STOR ' + file_remote, fp, bufsize)
# fp.close()

# data = {
#     'keywords': '宁波男科有问必答',
#     'url': 'https://m.120ask.com/askg/mip_detail/29325308',
#     'keywords_id': 12602,
#     'user_id': 45,
#     'rank': 1
# }
# # rand_str': '70082f1db65defc85072f83972d7fb84', 'timestamp': '1544750106540', 'user_id': 44
# import requests, json
# url = 'http://xiongzhanghao.zhugeyingxiao.com:8003/api/keyword_article_back_url/judgeLink?user_id=44&timestamp=1544750106540&rand_str=70082f1db65defc85072f83972d7fb84'
# ret= requests.post(url)
# json_data = ret.json().get('data')
# print(json_data)
# pcRequestHeader = [
#     'Mozilla/5.0 (Windows NT 5.1; rv:6.0.2) Gecko/20100101 Firefox/6.0.2',
#     'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_5) AppleWebKit/537.17 (KHTML, like Gecko) Chrome/24.0.1312.52 Safari/537.17',
#     'Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9.1.16) Gecko/20101130 Firefox/3.5.16',
#     'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.0; .NET CLR 1.1.4322)',
#     'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; WOW64; Trident/5.0)',
#     'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/45.0.2454.99 Safari/537.36',
#     'Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; .NET CLR 1.1.4322)',
#     'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 5.2)',
#     'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.13 (KHTML, like Gecko) Chrome/24.0.1290.1 Safari/537.13',
#     'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0)',
#     'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/45.0.2454.85 Safari/537.36',
#     'Mozilla/5.0 (Windows; U; Windows NT 5.2; zh-CN; rv:1.9.0.19) Gecko/2010031422 Firefox/3.0.19 (.NET CLR 3.5.30729)',
#     'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.2)',
#     'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.17 (KHTML, like Gecko) Chrome/24.0.1312.57 Safari/537.17',
#     'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/31.0.1650.63 Safari/537.36',
#     'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:24.0) Gecko/20100101 Firefox/24.0',
#     'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:2.0b13pre) Gecko/20110307 Firefox/4.0b13'
# ]


# from urllib import parse
# import random
# from requests.exceptions import ReadTimeout,  ConnectionError
#
# keywords = '宁波送子鸟男科医院官网'
#
# url = "http://www.baidu.com/s?wd={keywords}".format(
#     keywords=parse.quote(keywords)
# )
# print(url)
#
# headers = {
#     'User-Agent': pcRequestHeader[random.randint(0, len(pcRequestHeader) - 1)],
# }
# req_obj = requests.session()
# while True:
#     try:
#         try:
#             ret = req_obj.get(url, headers=headers, timeout=10)
#             break
#         except ReadTimeout:
#             time.sleep(1)
#     except ConnectionError:
#         time.sleep(1)
#
# page_source = ret.text
# soup = BeautifulSoup(page_source, 'lxml')
#
# for rank_num in range(1, 11):
#
#     div_tag = soup.find('div', id=str(rank_num), class_='result')
#     if div_tag:
#         f13 = div_tag.find('div', class_='f13')
#         panduan_url = f13.find('a').get_text()
#         if 'http://' not in panduan_url:
#             panduan_url = 'http://' + panduan_url
#         url_list = [
#             'http://4g.scgcyy.com',
#             'http://m.chyy120.com',
#             'http://m.szwk120.com',
#             'http://3g.ynttb.net/index.html',
#             'http://www.zjsznnk.com',
#             'http://www.zjnbsznfk120.com',
#             'http://www.bjwletyy.com',
#             'http://m.gzgbyy120.com',
#             'http://wap.tysgmr.com',
#             'http://m.oy120.com',
#             'http://5g.dzfyyy.com',
#             'http://m.glamzx.com',
#             'http://3g.meilianchen.cn',
#             'http://m.evercarebj.com',
#             'http://xzh.nk-hospital.mobi',
#             'http://m.28552855.com'
#         ]
#         flag = False
#         for i in url_list:
#             print(i, 'panduan_url--> ',panduan_url)
#             if i in panduan_url:
#                 flag = True
#         if flag:
#             url = div_tag.find('a').attrs.get("href")
#             while True:
#                 try:
#                     try:
#                         ret = requests.get(url, headers=headers, timeout=30)
#                         break
#                     except ReadTimeout:
#                         time.sleep(1)
#                 except ConnectionError:
#                     time.sleep(1)
#             url = ret.url
#             print('rank_num------------------------------------------------> ',rank_num , url)










